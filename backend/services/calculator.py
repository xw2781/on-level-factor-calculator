from __future__ import annotations

import datetime as dt
from dataclasses import dataclass

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from backend.models import (
    EffectiveDateRange,
    FormulaLine,
    InforceDateItem,
    InforceResponse,
    PlotPayload,
    QuarterRequest,
    WeightResponse,
    WorkbookSnapshotResponse,
)
from backend.services.clipboard import copy_text_to_clipboard
from backend.services.excel_session import ExcelSessionService


DATE_SCAN_LIMIT = 50
NO_PRIOR_RATE_CHANGE = "empty cell (no prior rate change)"


@dataclass(slots=True)
class QuarterBounds:
    year: int
    quarter: int
    start_date: dt.datetime
    end_date: dt.datetime
    start_display: str
    end_display: str
    start_month: int
    start_day: int
    end_month: int
    end_day: int


@dataclass(slots=True)
class SheetSnapshot:
    workbook_path: str
    sheet_name: str
    anchor_row: int
    anchor_column: int
    first_date_row_offset: int
    effective_dates: list[dt.datetime]
    policy_end_dates: list[dt.datetime]
    start_cell_refs: list[str]
    end_cell_refs: list[str]

    @property
    def anchor_cell(self) -> str:
        return f"{get_column_letter(self.anchor_column)}{self.anchor_row}"


@dataclass(slots=True)
class WeightContext:
    snapshot: SheetSnapshot
    request: QuarterRequest
    quarter: QuarterBounds
    inforce_start_dates: list[dt.datetime]
    inforce_end_dates: list[dt.datetime]
    inforce_start_refs: list[str]
    inforce_end_refs: list[str]
    quarter_start_ref: str
    quarter_end_ref: str
    quarter_weight_formula: str
    quarter_weight_value: float


class CalculatorError(RuntimeError):
    """Raised when the workbook selection cannot be converted into results."""


def parse_date_value(raw_value: object) -> dt.datetime | None:
    if isinstance(raw_value, dt.datetime):
        return raw_value
    if isinstance(raw_value, dt.date):
        return dt.datetime(raw_value.year, raw_value.month, raw_value.day)
    if isinstance(raw_value, str):
        cleaned_value = raw_value.strip()
        if "/" not in cleaned_value:
            return None

        date_token = cleaned_value.split()[0]
        date_parts = date_token.split("/")
        if len(date_parts) != 3:
            return None

        month_text, day_text, year_text = date_parts
        try:
            return dt.datetime(int(year_text), int(month_text), int(day_text))
        except ValueError:
            return None
    return None


def format_slash_date(value: dt.datetime) -> str:
    return f"{value.month}/{value.day}/{value.year}"


def format_iso_date(value: dt.datetime) -> str:
    return value.strftime("%Y-%m-%d")


def years_for_ui() -> list[int]:
    current_year = dt.date.today().year
    return list(range(current_year, 2018, -1))


def policy_end_dates(effective_dates: list[dt.datetime], policy_term_months: int) -> list[dt.datetime]:
    return [effective_date + relativedelta(months=+policy_term_months) for effective_date in effective_dates]


def quarter_bounds(year: int, quarter: int) -> QuarterBounds:
    if quarter == 1:
        return QuarterBounds(year, quarter, dt.datetime(year - 1, 12, 31), dt.datetime(year, 3, 31), f"12/31/{year - 1}", f"3/31/{year}", 12, 31, 3, 31)
    if quarter == 2:
        return QuarterBounds(year, quarter, dt.datetime(year, 3, 31), dt.datetime(year, 6, 30), f"3/31/{year}", f"6/30/{year}", 3, 31, 6, 30)
    if quarter == 3:
        return QuarterBounds(year, quarter, dt.datetime(year, 6, 30), dt.datetime(year, 9, 30), f"6/30/{year}", f"9/30/{year}", 6, 30, 9, 30)
    return QuarterBounds(year, quarter, dt.datetime(year, 9, 30), dt.datetime(year, 12, 31), f"9/30/{year}", f"12/31/{year}", 9, 30, 12, 31)


class CalculatorService:
    def __init__(self, excel_session: ExcelSessionService | None = None) -> None:
        self._excel_session = excel_session or ExcelSessionService()

    def app_options(self) -> dict[str, object]:
        return {
            "years": years_for_ui(),
            "quarters": [
                {"value": 1, "label": "Q1"},
                {"value": 2, "label": "Q2"},
                {"value": 3, "label": "Q3"},
                {"value": 4, "label": "Q4"},
            ],
            "policy_terms": [
                {"value": 12, "label": "12 mon policy"},
                {"value": 6, "label": "6 mon policy"},
            ],
        }

    def inspect_active_workbook(self, policy_term_months: int) -> WorkbookSnapshotResponse:
        snapshot = self._load_snapshot(policy_term_months)
        ranges = [
            EffectiveDateRange(index=index, start_date=format_slash_date(start_date), end_date=format_slash_date(end_date))
            for index, (start_date, end_date) in enumerate(
                zip(snapshot.effective_dates, snapshot.policy_end_dates),
                start=1,
            )
        ]
        return WorkbookSnapshotResponse(
            workbook_path=snapshot.workbook_path,
            sheet_name=snapshot.sheet_name,
            anchor_cell=snapshot.anchor_cell,
            effective_date_ranges=ranges,
        )

    def calculate_inforce_dates(self, request: QuarterRequest) -> InforceResponse:
        snapshot = self._load_snapshot(request.policy_term_months)
        inforce_dates, _ = self._inforce_dates_for_quarter(
            snapshot.effective_dates,
            snapshot.policy_end_dates,
            request.year,
            request.quarter,
            include_empty_marker=True,
        )

        items: list[InforceDateItem] = []
        for index, value in enumerate(inforce_dates, start=1):
            if isinstance(value, str):
                items.append(InforceDateItem(index=index, label=value, iso_date=None))
            else:
                items.append(
                    InforceDateItem(
                        index=index,
                        label=value.strftime("%Y-%m-%d"),
                        iso_date=format_iso_date(value),
                    )
                )

        return InforceResponse(
            selection_label=f"{request.year} Q{request.quarter} - {request.policy_term_months} mon policy",
            inforce_dates=items,
        )

    def calculate_weights(self, request: QuarterRequest) -> WeightResponse:
        context = self._build_weight_context(request)
        formula_strings, numeric_values = self._generate_area_formulas(context)
        clipboard_text = "\n".join(formula_strings)
        copied_to_clipboard = copy_text_to_clipboard(clipboard_text) if clipboard_text else False

        formula_lines = [
            FormulaLine(index=index, formula=formula, numeric_value=value)
            for index, (formula, value) in enumerate(zip(formula_strings, numeric_values), start=1)
        ]

        return WeightResponse(
            selection_label=f"{request.year} Q{request.quarter} - {request.policy_term_months} mon policy",
            quarter_weight_formula=context.quarter_weight_formula,
            quarter_weight_value=context.quarter_weight_value,
            formula_lines=formula_lines,
            clipboard_text=clipboard_text,
            copied_to_clipboard=copied_to_clipboard,
            plot=PlotPayload(
                year=request.year,
                quarter=request.quarter,
                policy_term_months=request.policy_term_months,
                inforce_start_dates=[format_iso_date(date_value) for date_value in context.inforce_start_dates],
                weight_values=numeric_values,
            ),
        )

    def _build_weight_context(self, request: QuarterRequest) -> WeightContext:
        snapshot = self._load_snapshot(request.policy_term_months)
        quarter = quarter_bounds(request.year, request.quarter)

        start_index, end_index, _ = self._inforce_indices(
            snapshot.effective_dates,
            snapshot.policy_end_dates,
            quarter,
        )

        inforce_start_dates = snapshot.effective_dates[start_index : end_index + 1]
        inforce_end_dates = policy_end_dates(inforce_start_dates, request.policy_term_months)
        inforce_start_refs = snapshot.start_cell_refs[start_index : end_index + 1]
        inforce_end_refs = snapshot.end_cell_refs[start_index : end_index + 1]

        quarter_start_ref = self._resolve_quarter_boundary_ref(snapshot, quarter.start_date, quarter.start_display)
        quarter_end_ref = self._resolve_quarter_boundary_ref(snapshot, quarter.end_date, quarter.end_display)

        quarter_weight_formula = f"(({quarter_end_ref}-{quarter_start_ref})/365)"
        if request.quarter == 1:
            weight_value = (dt.datetime(request.year, quarter.end_month, quarter.end_day) - dt.datetime(request.year - 1, quarter.start_month, quarter.start_day)).days / 365
        else:
            weight_value = (dt.datetime(request.year, quarter.end_month, quarter.end_day) - dt.datetime(request.year, quarter.start_month, quarter.start_day)).days / 365

        if not inforce_end_dates:
            raise CalculatorError("No in-force dates were found for the selected quarter.")

        if inforce_end_dates[0] > quarter.start_date:
            sentinel = dt.datetime(request.year - 1, 1, 1)
            inforce_start_dates.insert(0, sentinel)
            inforce_end_dates.insert(0, sentinel)
            inforce_start_refs.insert(0, f"{request.year - 1}/1/1")
            inforce_end_refs.insert(0, f"{request.year - 1}/1/1")

        return WeightContext(
            snapshot=snapshot,
            request=request,
            quarter=quarter,
            inforce_start_dates=inforce_start_dates,
            inforce_end_dates=inforce_end_dates,
            inforce_start_refs=inforce_start_refs,
            inforce_end_refs=inforce_end_refs,
            quarter_start_ref=quarter_start_ref,
            quarter_end_ref=quarter_end_ref,
            quarter_weight_formula=quarter_weight_formula,
            quarter_weight_value=weight_value,
        )

    def _load_snapshot(self, policy_term_months: int) -> SheetSnapshot:
        selection = self._excel_session.get_active_selection()
        try:
            workbook = load_workbook(selection.workbook_path, data_only=True)
        except FileNotFoundError as error:
            raise CalculatorError("The active workbook file could not be found on disk.") from error
        except PermissionError as error:
            raise CalculatorError("The active workbook could not be opened. Check that you have permission to read it.") from error
        except Exception as error:
            raise CalculatorError("The active workbook could not be read. Save the file and try again.") from error

        try:
            worksheet = workbook[selection.sheet_name]
        except KeyError as error:
            raise CalculatorError("The active Excel sheet could not be located in the workbook.") from error

        effective_dates, first_date_row_offset = self._extract_effective_dates(
            worksheet,
            selection.active_row,
            selection.active_column,
        )
        if not effective_dates:
            raise CalculatorError("Date header not found in worksheet. Select the header cell above the effective dates and try again.")

        start_cell_refs = [
            f"{get_column_letter(selection.active_column)}{selection.active_row + index + 1 + first_date_row_offset}"
            for index in range(len(effective_dates))
        ]

        date_lookup = self._build_date_lookup(worksheet)
        end_dates = policy_end_dates(effective_dates, policy_term_months)
        end_cell_refs: list[str] = []
        for start_ref, end_date in zip(start_cell_refs, end_dates):
            direct_ref = date_lookup.get(format_slash_date(end_date))
            if direct_ref is None:
                direct_ref = f"DATE(YEAR({start_ref}),MONTH({start_ref})+{policy_term_months},DAY({start_ref}))"
            end_cell_refs.append(direct_ref)

        return SheetSnapshot(
            workbook_path=selection.workbook_path,
            sheet_name=selection.sheet_name,
            anchor_row=selection.active_row,
            anchor_column=selection.active_column,
            first_date_row_offset=first_date_row_offset,
            effective_dates=effective_dates,
            policy_end_dates=end_dates,
            start_cell_refs=start_cell_refs,
            end_cell_refs=end_cell_refs,
        )

    def _extract_effective_dates(self, worksheet, anchor_row: int, anchor_column: int) -> tuple[list[dt.datetime], int]:
        dates: list[dt.datetime] = []
        first_date_row_offset = 0
        current_row = anchor_row
        initial_anchor_row = anchor_row

        for _ in range(DATE_SCAN_LIMIT):
            cell_value = worksheet.cell(current_row + 1, anchor_column).value
            parsed = parse_date_value(cell_value)
            if parsed is None:
                if dates:
                    break
            else:
                dates.append(parsed)
                if len(dates) == 1:
                    first_date_row_offset = current_row - initial_anchor_row
            current_row += 1

        return dates, first_date_row_offset

    def _build_date_lookup(self, worksheet) -> dict[str, str]:
        lookup: dict[str, str] = {}
        for row in worksheet.iter_rows():
            for cell in row:
                parsed = parse_date_value(cell.value)
                if parsed is None:
                    continue
                lookup.setdefault(format_slash_date(parsed), cell.coordinate)
        return lookup

    def _inforce_dates_for_quarter(
        self,
        effective_dates: list[dt.datetime],
        end_dates: list[dt.datetime],
        year: int,
        quarter_number: int,
        include_empty_marker: bool,
    ) -> tuple[list[dt.datetime | str], bool]:
        bounds = quarter_bounds(year, quarter_number)
        start_index, end_index, needs_empty_marker = self._inforce_indices(effective_dates, end_dates, bounds)
        dates: list[dt.datetime | str] = effective_dates[start_index : end_index + 1]
        if include_empty_marker and needs_empty_marker:
            dates.insert(0, NO_PRIOR_RATE_CHANGE)
        return dates, needs_empty_marker

    def _inforce_indices(
        self,
        effective_dates: list[dt.datetime],
        end_dates: list[dt.datetime],
        bounds: QuarterBounds,
    ) -> tuple[int, int, bool]:
        start_index = -100
        needs_empty_marker = False

        for index, policy_end_date in enumerate(end_dates):
            if policy_end_date > bounds.start_date and index != 0:
                start_index = index - 1
                break
            if policy_end_date > bounds.start_date and index == 0:
                start_index = 0
                needs_empty_marker = True
                break

        if start_index == -100:
            start_index = len(effective_dates) - 1

        end_index = len(effective_dates) - 1
        for index, effective_date in enumerate(effective_dates):
            if effective_date >= bounds.end_date:
                end_index = index - 1
                break

        return start_index, end_index, needs_empty_marker

    def _resolve_quarter_boundary_ref(self, snapshot: SheetSnapshot, boundary_date: dt.datetime, boundary_text: str) -> str:
        try:
            workbook = load_workbook(snapshot.workbook_path, data_only=True)
            worksheet = workbook[snapshot.sheet_name]
        except Exception:
            return f"DATE({boundary_date.year},{boundary_date.month},{boundary_date.day})"

        target_date = parse_date_value(boundary_text)
        for row_index in range(1, worksheet.max_row + 1):
            for column_index in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row_index, column_index).value
                parsed = parse_date_value(cell_value)
                if parsed == target_date or cell_value == boundary_text:
                    return f"{get_column_letter(column_index)}{row_index}"

        return f"DATE({boundary_date.year},{boundary_date.month},{boundary_date.day})"

    def _generate_area_formulas(self, context: WeightContext) -> tuple[list[str], list[float]]:
        request = context.request
        quarter = context.quarter
        square_factor = 0.5 if request.policy_term_months == 12 else 1.0
        width_factor = 1 if request.policy_term_months == 12 else 2

        formula_strings: list[str] = []
        numeric_values: list[float] = []

        def squared_formula(expression: str) -> str:
            if square_factor == 0.5:
                return f"0.5*(({expression})/365)^2"
            return f"(({expression})/365)^2"

        def squared_value(date_delta: dt.timedelta) -> float:
            return square_factor * ((date_delta / dt.timedelta(days=1)) / 365) ** 2

        def trapezoid_formula(end_ref: str) -> str:
            if width_factor == 1:
                return f"(({end_ref}-{context.quarter_end_ref})/365+({end_ref}-{context.quarter_start_ref})/365)*{context.quarter_weight_formula}*0.5"
            return f"(2*({end_ref}-{context.quarter_end_ref})/365+2*({end_ref}-{context.quarter_start_ref})/365)*{context.quarter_weight_formula}*0.5"

        def trapezoid_value(end_date: dt.datetime) -> float:
            return (
                (
                    width_factor * (end_date - quarter.end_date) / dt.timedelta(days=1) / 365
                    + width_factor * (end_date - quarter.start_date) / dt.timedelta(days=1) / 365
                )
                * context.quarter_weight_value
                * 0.5
            )

        def band_formula(next_start_ref: str, previous_start_ref: str) -> str:
            if width_factor == 1:
                return f"(({next_start_ref}-{previous_start_ref})/365)*{context.quarter_weight_formula}"
            return f"2*(({next_start_ref}-{previous_start_ref})/365)*{context.quarter_weight_formula}"

        def band_value(next_start_date: dt.datetime, previous_start_date: dt.datetime) -> float:
            return width_factor * ((next_start_date - previous_start_date) / dt.timedelta(days=1) / 365) * context.quarter_weight_value

        if not context.inforce_end_dates:
            raise CalculatorError("No in-force dates were found for the selected quarter.")

        for index in range(1, len(context.inforce_start_dates) + 1):
            previous_end_date = context.inforce_end_dates[index - 1]
            previous_start_date = context.inforce_start_dates[index - 1]
            previous_end_ref = context.inforce_end_refs[index - 1]
            previous_start_ref = context.inforce_start_refs[index - 1]

            if index == len(context.inforce_start_dates):
                if previous_end_date < quarter.end_date:
                    if previous_end_date <= quarter.start_date:
                        formula_strings.append(f"=1*{context.quarter_weight_formula}")
                        numeric_values.append(context.quarter_weight_value)
                    else:
                        formula_strings.append(f"=1*{context.quarter_weight_formula}-{squared_formula(f'{previous_end_ref}-{context.quarter_start_ref}')}")
                        numeric_values.append(context.quarter_weight_value - squared_value(previous_end_date - quarter.start_date))

                if previous_end_date >= quarter.end_date and previous_start_date < quarter.start_date:
                    formula_strings.append(f"={trapezoid_formula(previous_start_ref)}")
                    numeric_values.append(
                        (
                            (
                                width_factor * (quarter.start_date - previous_start_date) / dt.timedelta(days=1) / 365
                                + width_factor * (quarter.end_date - previous_start_date) / dt.timedelta(days=1) / 365
                            )
                            * context.quarter_weight_value
                            * 0.5
                        )
                    )

                if previous_start_date >= quarter.start_date:
                    formula_strings.append(f"={squared_formula(f'{context.quarter_end_ref}-{previous_start_ref}')}")
                    numeric_values.append(squared_value(quarter.end_date - previous_start_date))
                continue

            next_end_date = context.inforce_end_dates[index]
            next_start_date = context.inforce_start_dates[index]
            next_end_ref = context.inforce_end_refs[index]
            next_start_ref = context.inforce_start_refs[index]

            if previous_end_date <= quarter.start_date:
                if next_end_date <= quarter.end_date:
                    formula_strings.append(f"={squared_formula(f'{next_end_ref}-{context.quarter_start_ref}')}")
                    numeric_values.append(squared_value(next_end_date - quarter.start_date))

                if next_end_date > quarter.end_date and next_start_date <= quarter.start_date:
                    formula_strings.append(f"={trapezoid_formula(next_end_ref)}")
                    numeric_values.append(trapezoid_value(next_end_date))

                if next_start_date > quarter.start_date:
                    formula_strings.append(f"=1*{context.quarter_weight_formula}-{squared_formula(f'{context.quarter_end_ref}-{next_start_ref}')}")
                    numeric_values.append(context.quarter_weight_value - squared_value(quarter.end_date - next_start_date))
                continue

            if previous_end_date > quarter.start_date and previous_end_date <= quarter.end_date:
                if next_end_date <= quarter.end_date:
                    formula_strings.append(
                        f"={squared_formula(f'{next_end_ref}-{context.quarter_start_ref}')}-{squared_formula(f'{previous_end_ref}-{context.quarter_start_ref}')}"
                    )
                    numeric_values.append(
                        squared_value(next_end_date - quarter.start_date)
                        - squared_value(previous_end_date - quarter.start_date)
                    )

                if next_end_date > quarter.end_date and next_start_date <= quarter.start_date:
                    formula_strings.append(
                        f"={trapezoid_formula(next_end_ref)}-{squared_formula(f'{previous_end_ref}-{context.quarter_start_ref}')}"
                    )
                    numeric_values.append(
                        trapezoid_value(next_end_date)
                        - squared_value(previous_end_date - quarter.start_date)
                    )

                if next_start_date > quarter.start_date:
                    formula_strings.append(
                        f"=1*{context.quarter_weight_formula}-{squared_formula(f'{previous_end_ref}-{context.quarter_start_ref}')}-{squared_formula(f'{context.quarter_end_ref}-{next_start_ref}')}"
                    )
                    numeric_values.append(
                        context.quarter_weight_value
                        - squared_value(previous_end_date - quarter.start_date)
                        - squared_value(quarter.end_date - next_start_date)
                    )
                continue

            if previous_end_date > quarter.end_date and previous_start_date <= quarter.start_date:
                if next_start_date <= quarter.start_date:
                    formula_strings.append(f"={band_formula(next_start_ref, previous_start_ref)}")
                    numeric_values.append(band_value(next_start_date, previous_start_date))

                if next_start_date > quarter.start_date:
                    formula_strings.append(
                        f"=1*{context.quarter_weight_formula}-{trapezoid_formula(previous_end_ref)}-{squared_formula(f'{context.quarter_end_ref}-{next_start_ref}')}"
                    )
                    numeric_values.append(
                        context.quarter_weight_value
                        - trapezoid_value(previous_end_date)
                        - squared_value(quarter.end_date - next_start_date)
                    )
                continue

            if previous_start_date > quarter.start_date:
                formula_strings.append(
                    f"={squared_formula(f'{context.quarter_end_ref}-{previous_start_ref}')}-{squared_formula(f'{context.quarter_end_ref}-{next_start_ref}')}"
                )
                numeric_values.append(
                    squared_value(quarter.end_date - previous_start_date)
                    - squared_value(quarter.end_date - next_start_date)
                )

        return formula_strings, numeric_values
